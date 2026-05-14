"""
data_processor.py
-----------------
Loads, cleans, and joins employee + timesheet data.
Produces a final reminder_df of employees who need reminders.

Folder structure expected:
    data/
        employee_data_Export.csv
        Detail_Timesheet_Report_-_LiveReport.xlsx
"""

import os
import re
import datetime
import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv

# ─────────────────────────────────────────────
# CONFIG — update paths here if folder changes
# ─────────────────────────────────────────────
DATA_FOLDER    = "data"
EMPLOYEE_FILE  = os.path.join(DATA_FOLDER, "employee_data_Export.csv")
TIMESHEET_FILE = os.path.join(DATA_FOLDER, "Timesheet_Report.xlsx")


# ═══════════════════════════════════════════════════════════
# HELPER — NAME CLEANING
# ═══════════════════════════════════════════════════════════

def clean_name(name: str) -> str:
    """
    Standardizes a name string:
    - Strips leading/trailing whitespace
    - Collapses multiple spaces into one
    - Title cases the result
    """
    if not name:
        return ''
    name = name.strip()
    name = re.sub(r'\s+', ' ', name)
    name = name.title()
    return name


# ═══════════════════════════════════════════════════════════
# STEP 1 — LOAD & CLEAN EMPLOYEE DATA
# ═══════════════════════════════════════════════════════════

def load_employee_data() -> pd.DataFrame:
    """
    Loads employee CSV and keeps only columns relevant to reminders.
    Returns a clean employee DataFrame.
    """
    df = pd.read_csv(EMPLOYEE_FILE, encoding='latin1')

    df = df[[
        'AccountEmployeeId',
        'FirstName',
        'LastName',
        'EMailAddress',
        'MobilePhoneNo',
        'JobTitle',
        'EmployeeManager',
        'AccountEmployeeType',
    ]].copy()

    df.rename(columns={
        'AccountEmployeeId'  : 'employee_id',
        'FirstName'          : 'first_name',
        'LastName'           : 'last_name',
        'EMailAddress'       : 'email',
        'MobilePhoneNo'      : 'phone',
        'JobTitle'           : 'job_title',
        'EmployeeManager'    : 'manager_name',
        'AccountEmployeeType': 'employee_type',
    }, inplace=True)

    df['first_name'] = df['first_name'].str.strip()
    df['last_name']  = df['last_name'].str.strip()
    df['email']      = df['email'].str.strip().str.lower()
    df['email'] = df['email'].str.replace('narwalinc.com', 'narwal.ai', regex=False)

    # Build normalised full_name join key
    df['full_name'] = (df['first_name'] + ' ' + df['last_name']).apply(clean_name)

    # Remove disabled employees
    disabled_mask = df['full_name'].str.contains(r'\(Disabled\)', case=False, na=False)
    df = df[~disabled_mask].copy()

    print(f"[Employee Data]  Loaded {len(df)} active employees | "
          f"{df['email'].notna().sum()} emails | "
          f"{df['phone'].notna().sum()} phone numbers")

    return df


# ═══════════════════════════════════════════════════════════
# STEP 2 — LOAD & PARSE DETAIL TIMESHEET
# ═══════════════════════════════════════════════════════════

def load_timesheet_data() -> pd.DataFrame:
    """
    Parses the Detail Timesheet Report xlsx.
    Returns a daily DataFrame — one row per employee per day.
    """
    wb = load_workbook(TIMESHEET_FILE, read_only=True)
    ws = wb.active

    records = []
    current_employee = None

    for row in ws.iter_rows(values_only=True):
        # Employee name row: col[0] has name, col[1] is None
        if (row[0]
                and row[1] is None
                and row[0] not in ('Detail Timesheet Report', 'Employee Name')
                and str(row[0]).strip() not in ('', 'Full Summary')
        ):
            raw_name = str(row[0])
            current_employee = None if '(Disabled)' in raw_name else clean_name(raw_name)

        # Daily entry row: col[1] is a date
        elif current_employee and isinstance(row[1], datetime.datetime):
            records.append({
                'full_name'      : current_employee,
                'date'           : row[1].date(),
                'client_name'    : row[2],
                'project_name'   : row[3],
                'hours'          : row[6] or 0,
                'approval_status': row[7],
            })

    df = pd.DataFrame(records)

    print(f"[Timesheet Data] Loaded {len(df)} daily records | "
          f"{df['full_name'].nunique()} unique employees | "
          f"Date range: {df['date'].min()} → {df['date'].max()}")

    return df


# ═══════════════════════════════════════════════════════════
# STEP 3 — BUILD REMINDER DATAFRAME
# ═══════════════════════════════════════════════════════════

def build_reminder_df(
    employee_df  : pd.DataFrame,
    timesheet_df : pd.DataFrame,
) -> pd.DataFrame:
    """
    Filters employees who have any day with 'Not Submitted' status.
    Aggregates all not-submitted dates into a single row per employee.
    Joins their contact details from the Employee Master.
    Returns reminder_df — one row per employee who needs a reminder.
    """

    # Filter to Not Submitted days only
    not_submitted = timesheet_df[
        timesheet_df['approval_status'] == 'Not Submitted'
    ].copy()

    not_submitted = not_submitted.sort_values(['full_name', 'date'])

    print(f"[Not Submitted]  {not_submitted['full_name'].nunique()} employees | "
          f"{len(not_submitted)} not-submitted days")

    # Group by employee — one row per employee with all dates aggregated
    aggregated = (
        not_submitted
        .groupby('full_name')
        .agg(
            not_submitted_dates = ('date', lambda x: ', '.join(d.strftime('%d %b %Y') for d in sorted(x))),
            not_submitted_count = ('date', 'count'),
            client_name         = ('client_name', 'first'),
            project_name        = ('project_name', 'first'),
        )
        .reset_index()
    )

    # Join employee contact details
    reminder_df = aggregated.merge(
        employee_df,
        on='full_name',
        how='left',
    )

    # Warn and drop any unmatched
    unmatched = reminder_df['email'].isna().sum()
    if unmatched:
        print(f"[Warning]        {unmatched} employees not matched in Employee Master — skipped")
        for name in reminder_df[reminder_df['email'].isna()]['full_name']:
            print(f"                 • {name}")
    reminder_df = reminder_df.dropna(subset=['email']).copy()

    reminder_df.reset_index(drop=True, inplace=True)

    print(f"[Reminder List]  {len(reminder_df)} employees | "
          f"{reminder_df['email'].notna().sum()} with email | "
          f"{reminder_df['phone'].notna().sum()} with phone")

    return reminder_df


# ═══════════════════════════════════════════════════════════
# MAIN — RUN AS SCRIPT TO TEST
# ═══════════════════════════════════════════════════════════

if __name__ == "__main__":

    print("=" * 60)
    print("  TIMESHEET REMINDER — DATA PROCESSOR")
    print("=" * 60)

    # Step 1 — Load employees
    employee_df  = load_employee_data()

    # Step 2 — Load timesheet
    timesheet_df = load_timesheet_data()

    # Step 3 — Build reminder list
    reminder_df  = build_reminder_df(employee_df, timesheet_df)

    # Step 4 — Save to Excel
    output_path = os.path.join(DATA_FOLDER, "reminder_list.xlsx")
    reminder_df.to_excel(output_path, index=False)
    print(f"\nReminder list saved to {output_path}")

    # ── Summary ──
    print()
    print("=" * 60)
    print("  REMINDER LIST PREVIEW")
    print("=" * 60)
    print(reminder_df[[
        'full_name', 'email', 'phone',
        'manager_name', 'not_submitted_dates',
        'not_submitted_count', 'client_name', 'project_name',
    ]].to_string(index=False))

    print()
    print(f"  Total employees to remind : {len(reminder_df)}")
    print(f"  Employees with email      : {reminder_df['email'].notna().sum()}")
    print(f"  Employees with phone      : {reminder_df['phone'].notna().sum()}")
    print("=" * 60)